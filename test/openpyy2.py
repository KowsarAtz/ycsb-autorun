from openpyxl import load_workbook

wb = load_workbook(filename = "./sample.xlsx")
ws = wb.active

ws['C1'] = "C"

wb.save("./sample.xlsx")